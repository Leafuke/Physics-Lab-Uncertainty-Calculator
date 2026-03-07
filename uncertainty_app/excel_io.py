from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .calculations import CalculationResult, format_number, normalize_decimal_places, rounded_measurement
from .models import ProjectData, b_source_display_name

KNOWN_VALUE_HEADERS = {
    "测量值",
    "测量数据",
    "原始数据",
    "数据",
    "value",
    "values",
    "measurement",
    "measurements",
    "x",
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="EFF4FB")
HEADER_FONT = Font(bold=True, color="243B53")
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def import_measurements_from_excel(file_path: str) -> tuple[list[float], str]:
    pd = _import_pandas()
    workbook = pd.ExcelFile(file_path)

    for sheet_name in workbook.sheet_names:
        dataframe = pd.read_excel(file_path, sheet_name=sheet_name)
        values = _extract_measurement_values(dataframe, pd)
        if values:
            return values, sheet_name

    raise ValueError("未找到可用的测量值列。请使用带有“测量值”列名的模板，或确保至少有一列数值数据。")


def export_project_to_excel(project: ProjectData, result: CalculationResult, file_path: str) -> None:
    target = Path(file_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = _create_workbook("结果导出")
    result_number_format = _excel_number_format(project.result_decimal_places)
    unit_suffix = f" {project.unit}" if project.unit else ""
    rounded_value, rounded_uncertainty = rounded_measurement(
        result.mean,
        result.expanded_uncertainty,
        project.result_decimal_places,
    )

    overview_sheet = workbook.active
    overview_sheet.title = "项目概览"
    overview_rows = [
        ("测量量", project.quantity_name or "未命名", None),
        ("单位", project.unit or "-", None),
        ("测量次数", result.sample_count, "0"),
        ("平均值", result.mean, result_number_format),
        ("A类标准不确定度", result.type_a_uncertainty, result_number_format),
        ("B类合成标准不确定度", result.type_b_uncertainty, result_number_format),
        ("合成标准不确定度", result.combined_uncertainty, result_number_format),
        ("扩展不确定度", result.expanded_uncertainty, result_number_format),
        ("覆盖因子 k", result.coverage_factor, "0.###"),
        (
            "结果保留小数位",
            "自动修约" if normalize_decimal_places(project.result_decimal_places) is None else f"{project.result_decimal_places} 位",
            None,
        ),
        (
            "最终结果文本",
            f"{project.quantity_name or '测量量'} = ({rounded_value} ± {rounded_uncertainty}){unit_suffix}, k = {format_number(result.coverage_factor)}",
            None,
        ),
    ]
    _write_key_value_sheet(overview_sheet, overview_rows)

    raw_values_sheet = workbook.create_sheet("原始数据")
    _write_table_sheet(
        raw_values_sheet,
        ["序号", "测量值"],
        [[index, value] for index, value in enumerate(project.measured_values, start=1)],
        numeric_formats={1: "0", 2: _excel_number_format(None)},
    )

    b_sources_sheet = workbook.create_sheet("B类评定")
    _write_table_sheet(
        b_sources_sheet,
        ["名称", "类型", "输入值", "分布因子", "标准不确定度", "说明"],
        [
            [
                component.name,
                b_source_display_name(component.source_type),
                component.input_value,
                component.divisor,
                component.standard_uncertainty,
                component.note,
            ]
            for component in result.b_components
        ],
        numeric_formats={3: _excel_number_format(None), 4: "0.############", 5: result_number_format},
    )

    process_sheet = workbook.create_sheet("计算过程")
    _write_table_sheet(
        process_sheet,
        ["类别", "项目", "数值", "说明"],
        [[row["类别"], row["项目"], row["数值"], row["说明"]] for row in result.process_rows],
    )

    report_sheet = workbook.create_sheet("文本报告")
    _write_table_sheet(report_sheet, ["文本报告"], [[line] for line in result.summary_text.splitlines()])
    workbook.save(target)


def export_data_to_excel(project: ProjectData, file_path: str) -> None:
    target = Path(file_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = _create_workbook("数据导出")
    info_sheet = workbook.active
    info_sheet.title = "项目数据"
    info_rows = [
        ("测量量", project.quantity_name or "未命名", None),
        ("单位", project.unit or "-", None),
        ("覆盖因子 k", project.coverage_factor, "0.###"),
        (
            "结果保留小数位",
            "自动修约" if normalize_decimal_places(project.result_decimal_places) is None else f"{project.result_decimal_places} 位",
            None,
        ),
        ("最近导入来源", project.last_import_path or "-", None),
        ("备注", project.notes or "", None),
    ]
    _write_key_value_sheet(info_sheet, info_rows)

    measurement_sheet = workbook.create_sheet("测量数据")
    _write_table_sheet(
        measurement_sheet,
        ["序号", "测量值"],
        [[index, value] for index, value in enumerate(project.measured_values, start=1)],
        numeric_formats={1: "0", 2: _excel_number_format(None)},
    )

    b_sheet = workbook.create_sheet("B类输入")
    _write_table_sheet(
        b_sheet,
        ["类型", "名称", "输入值", "分布因子", "备注"],
        [
            [
                b_source_display_name(source.source_type),
                source.name,
                source.value,
                source.normalized_divisor(),
                source.notes,
            ]
            for source in project.b_sources
        ],
        numeric_formats={3: _excel_number_format(None), 4: "0.############"},
    )

    workbook.save(target)


def _extract_measurement_values(dataframe: Any, pandas_module: Any) -> list[float]:
    if dataframe is None or dataframe.empty:
        return []

    for column_name in dataframe.columns:
        normalized = _normalize_header(column_name)
        if normalized in {_normalize_header(header) for header in KNOWN_VALUE_HEADERS}:
            values = _clean_numeric_series(dataframe[column_name], pandas_module)
            if values:
                return values

    candidates: list[tuple[int, list[float]]] = []
    for column_name in dataframe.columns:
        values = _clean_numeric_series(dataframe[column_name], pandas_module)
        if values:
            candidates.append((len(values), values))

    if not candidates:
        return []
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def _clean_numeric_series(series: Any, pandas_module: Any) -> list[float]:
    cleaned = pandas_module.to_numeric(series, errors="coerce").dropna().astype(float)
    return [float(value) for value in cleaned.tolist()]


def _normalize_header(header: Any) -> str:
    text = str(header).strip().lower()
    return "".join(character for character in text if character.isalnum())


def _import_pandas() -> Any:
    try:
        import pandas as pd
    except ImportError as exc:
        raise RuntimeError("需要先安装 pandas 和 openpyxl，才能导入或导出 Excel。") from exc

    return pd


def _create_workbook(subject: str) -> Workbook:
    workbook = Workbook()
    workbook.properties.creator = "Leafuke"
    workbook.properties.title = f"物理实验不确定度{subject}"
    workbook.properties.subject = subject
    return workbook


def _write_key_value_sheet(sheet: Any, rows: list[tuple[str, Any, str | None]]) -> None:
    sheet.append(["字段", "值"])
    _style_header_row(sheet, 1)
    for key, value, number_format in rows:
        sheet.append([key, value])
        value_cell = sheet.cell(sheet.max_row, 2)
        value_cell.alignment = BODY_ALIGNMENT
        if number_format:
            value_cell.number_format = number_format
    sheet.freeze_panes = "A2"
    _autosize_sheet_columns(sheet)


def _write_table_sheet(
    sheet: Any,
    headers: list[str],
    rows: list[list[Any]],
    numeric_formats: dict[int, str] | None = None,
) -> None:
    numeric_formats = numeric_formats or {}
    sheet.append(headers)
    _style_header_row(sheet, 1)
    if rows:
        for row_values in rows:
            sheet.append(row_values)
    else:
        sheet.append([""] * len(headers))

    for row_index in range(2, sheet.max_row + 1):
        for column_index in range(1, len(headers) + 1):
            cell = sheet.cell(row_index, column_index)
            cell.alignment = BODY_ALIGNMENT
            if column_index in numeric_formats and isinstance(cell.value, (int, float)):
                cell.number_format = numeric_formats[column_index]
                cell.alignment = CENTER_ALIGNMENT
    sheet.freeze_panes = "A2"
    _autosize_sheet_columns(sheet)


def _style_header_row(sheet: Any, row_index: int) -> None:
    for cell in sheet[row_index]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT


def _autosize_sheet_columns(sheet: Any) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 42)


def _excel_number_format(decimal_places: int | None) -> str:
    decimal_places = normalize_decimal_places(decimal_places)
    if decimal_places is None:
        return "0.############"
    if decimal_places == 0:
        return "0"
    return "0." + ("0" * decimal_places)